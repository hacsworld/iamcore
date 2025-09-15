# core/readers.py
from __future__ import annotations

import io
import csv
from typing import Tuple
from PIL import Image, ExifTags
import chardet
import pdfplumber
from docx import Document
from openpyxl import load_workbook

# ----- Optional OCR -----
_OCR_READY = False
try:
    import pytesseract  # type: ignore
    from shutil import which

    if which("tesseract"):
        _OCR_READY = True
except Exception:
    _OCR_READY = False


# ----- Helpers -----
def _decode_bytes(b: bytes) -> str:
    if not b:
        return ""
    enc = chardet.detect(b).get("encoding") or "utf-8"
    try:
        return b.decode(enc, errors="ignore")
    except Exception:
        return b.decode("utf-8", errors="ignore")


# ----- Readers -----
def read_txt(data: bytes) -> str:
    return _decode_bytes(data)


def read_csv(data: bytes) -> str:
    txt = _decode_bytes(data)
    out = []
    for row in csv.reader(io.StringIO(txt)):
        out.append(" | ".join(row))
    return "\n".join(out)


def read_pdf(data: bytes) -> str:
    out = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            out.append(page.extract_text() or "")
    return "\n".join(out).strip()


def read_docx(data: bytes) -> str:
    f = io.BytesIO(data)
    doc = Document(f)
    return "\n".join(p.text for p in doc.paragraphs)


def read_xlsx(data: bytes) -> str:
    out = []
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    for ws in wb.worksheets:
        out.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            out.append(" | ".join(cells))
    return "\n".join(out)


def read_image(data: bytes) -> str:
    buff = io.BytesIO(data)
    img = Image.open(buff)
    meta = [f"IMAGE: {img.format} {img.size[0]}x{img.size[1]} mode={img.mode}"]

    # EXIF (best-effort)
    try:
        exif = img.getexif()
        if exif:
            tagmap = {ExifTags.TAGS.get(k, k): v for k, v in exif.items() if isinstance(k, int)}
            useful = []
            for k in ("Model", "Make", "DateTimeOriginal", "LensModel"):
                if tagmap.get(k):
                    useful.append(f"{k}: {tagmap[k]}")
            if useful:
                meta.append("EXIF: " + "; ".join(useful))
    except Exception:
        pass

    if not _OCR_READY:
        meta.append("OCR: not available (install Tesseract to enable)")
        return "\n".join(meta)

    text = ""
    try:
        text = pytesseract.image_to_string(img)  # type: ignore
    except Exception:
        text = ""
    block = "\n".join(meta + (["OCR:\n" + text] if text else []))
    return block.strip()


# ----- Sniffer -----
def sniff_and_read(filename: str, data: bytes) -> Tuple[str, str]:
    """
    Вернёт (kind, text), где kind ∈ {txt,csv,pdf,docx,xlsx,image,unknown}.
    """
    lower = (filename or "").lower()

    if lower.endswith(".txt"):
        return "txt", read_txt(data)
    if lower.endswith(".csv"):
        return "csv", read_csv(data)
    if lower.endswith(".pdf"):
        return "pdf", read_pdf(data)
    if lower.endswith(".docx"):
        return "docx", read_docx(data)
    if lower.endswith(".xlsx"):
        return "xlsx", read_xlsx(data)
    if lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff")):
        return "image", read_image(data)

    # fallback: попробуем как текст
    txt = _decode_bytes(data)
    if txt and any(c in txt for c in ("\n", " ")):
        return "txt", txt

    return "unknown", ""
